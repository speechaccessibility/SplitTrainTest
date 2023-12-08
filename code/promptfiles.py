import openpyxl,json,argparse,os

################################################################################################
# Load all prompts

def load_prompts(promptsroot):
    '''
    Load the prompt files that specify which list a person is using.

    @param:
    promptsroot (str): directory that contains the prompt lists

    @return:
    prompts (dict):  
     prompts[listnum][block][i][0] is category
     prompts[listnum][block][i][1] is a subcategory, or empty string
     prompts[listnum][block][i][2] is a prompt text
    '''
    
    prompts = { listnum:{block:[] for block in range(1,11)} for listnum in range(1,11) }

    promptlists = [
        'commands_b2-9_DS _35perBlock.xlsx',
        'commands b2-9_PD.xlsx'
    ]
              
    # Files in which each sheet is named 'list%d'%(listnum)
    for filename in promptlists:
        pathname = os.path.join(promptsroot, filename)
        wb = openpyxl.load_workbook(pathname)
        for listnum in range(1,11):
            ws = wb['list%d'%(listnum)]
            for block in range(2,10):
                for ind in range(1,31):
                    x = ind + 30*(block-2)
                    prompts[listnum][block].append(["Digital Assistant Commands",
                                                    ws['B%d'%(x)].value,
                                                    ws['C%d'%(x)].value])
    return prompts
    
################################################################################################
# load_PD_prompts (deprecated)
#

def load_PD_prompts():
    '''
    Load PD prompt files (deprecated).
    
    @return:
    prompts (dict):  
     prompts[listnum][block][i][0] is category
     prompts[listnum][block][i][1] is a subcategory, or empty string
     prompts[listnum][block][i][2] is a prompt text

    novel_sentences: a list of 12000 ["Novel Sentences",'',prompt_text]
    '''
    prompts = { listnum:{block:{} for block in range(1,11)} for listnum in range(1,11) }
    
    with open('../lists/commands_b1_b10_PD.txt') as f:
        lines = f.readlines()
    for ind in range(1,31):
        prompts[1][10][ind] = prompts[1][1][ind] = [ "Digital Assistant Commands", '', lines[ind-1] ]
        for listnum in range(2,11):
            prompts[listnum][10][ind] = prompts[listnum][1][ind] = prompts[1][1][ind]
        
    wb = openpyxl.load_workbook('../lists/commands b2-9_PD.xlsx')
    for listnum in range(1,11):
        ws = wb['list%d'%(listnum)]
        for block in range(2,10):
            for ind in range(1,31):
                x = ind + 30*(block-2)
                prompts[listnum][block][ind]=["Digital Assistant Commands",
                                              ws['B%d'%(x)].value,
                                              ws['C%d'%(x)].value]
    
    ws = openpyxl.load_workbook('../lists/novel_b1_b10_PD.xlsx').active
    for ind in range(31,41):
        prompts[1][10][ind]=prompts[1][1][ind]=["Novel Sentences", '', ws['A%d'%(ind-30)].value ]
        for listnum in range(2,11):
            prompts[listnum][10][ind] = prompts[listnum][1][ind] = prompts[1][1][ind]
    
    ws = openpyxl.load_workbook('../lists/novel-b2-9-150_PD.xlsx').active
    novel_sentences = [ ["Novel Sentences",'',ws['B%d'%(d)].value] for d in range(1,12001) ]
    
    ws = openpyxl.load_workbook('../lists/Spontaneous speech prompts_PD.xlsx').active
    for block in range(1,11):
        for i in range(1,6):
            ind = 40+i
            rownum = 10+i+4*block
            prompts[1][block][ind]=[ "Spontaneous Speech Prompts",
                                     ws['A%d'%(ind-30)].value,
                                     ws['B%d'%(ind-30)].value ]
            for listnum in range(2,11):
                prompts[listnum][block][ind] = prompts[1][block][ind]

    return prompts, novel_sentences
    
################################################################################################
# Command line arguments
#
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description     = 'Load transcripts, and save them in two json files to test',
        formatter_class = argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('promptsfile',help = 'Directory containing unsplit dataset')
    parser.add_argument('novelfile',help = 'Directory containing unsplit dataset')
    args   = parser.parse_args()

    prompts, novel_sentences = load_prompts()
    with open(args.promptsfile,'w') as f:
        json.dump(prompts, f, indent=2)
    with open(args.novelfile,'w') as f:
        json.dump(novel_sentences,f, indent=2)
